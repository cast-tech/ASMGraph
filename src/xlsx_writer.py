import openpyxl
from openpyxl.styles import Font, Alignment, DEFAULT_FONT
from typing import List, Dict, NoReturn
from .graph import Node, FlowGraph


DEFAULT_FONT.name = "Arial"

rich_text_is_available = True
try:
    from openpyxl.cell.text import InlineFont
    from openpyxl.cell.rich_text import TextBlock, CellRichText

    red_font = InlineFont(rFont="Arial", b=True, color='FF0000')

except Exception as e:
    print("Rich text is not available, cannot highlight fuse cases.")
    rich_text_is_available = False


class XLSXWriter:
    def __init__(self, xlsx_file: str):
        self.__xlsx_file = xlsx_file
        self.__workbook = openpyxl.Workbook()
        self.__worksheet = self.__workbook.active
        self.__worksheet.__rows = []

    def create_asm_sheet(self) -> NoReturn:
        self.__worksheet.title = "AsmSheet"

        self.__worksheet["A1"] = "Function Name"
        self.__worksheet["B1"] = "Basic Block's address (Label)"
        self.__worksheet["C1"] = "ASM"
        self.__worksheet["D1"] = "Execution count"

        self.__worksheet.column_dimensions["A"].width = 30
        self.__worksheet.column_dimensions["B"].width = 30
        self.__worksheet.column_dimensions["C"].width = 70
        self.__worksheet.column_dimensions["D"].width = 15

        bold_font = Font(bold=True)
        for cell in ['A1', 'B1', 'C1', 'D1']:
            self.__worksheet[cell].font = bold_font

    def create_checkers_sheet(self, title: str) -> NoReturn:
        if title in self.__workbook.get_sheet_names():
            return

        if 'Sheet' in self.__workbook.get_sheet_names():
            self.__worksheet.title = title
        else:
            self.__workbook.create_sheet(title)
            self.__worksheet = self.__workbook.get_sheet_by_name(title)
            self.__worksheet.title = title

        self.__worksheet = self.__workbook.get_sheet_by_name(title)
        self.__worksheet.__rows = []

        self.__worksheet["A1"] = "Function Name"
        self.__worksheet["B1"] = "Basic Block"
        self.__worksheet["C1"] = "Fuse"
        self.__worksheet["D1"] = "BB Execution count"
        self.__worksheet["E1"] = "Instruction Profit"

        self.__worksheet.column_dimensions["A"].width = 30
        self.__worksheet.column_dimensions["B"].width = 70
        self.__worksheet.column_dimensions["C"].width = 70
        self.__worksheet.column_dimensions["D"].width = 20
        self.__worksheet.column_dimensions["E"].width = 20

        bold_font = Font(bold=True)
        for cell in ['A1', 'B1', 'C1', 'D1', 'E1']:
            self.__worksheet[cell].font = bold_font
            self.__worksheet[cell].alignment = Alignment(horizontal="center", vertical="top")

        self.__workbook.save(self.__xlsx_file)

    def append_checker_result(self, title: str,
                              func_name: str,
                              node: Node,
                              fusions: List[Dict],
                              highlight_fuse=False) -> NoReturn:

        self.create_checkers_sheet(title)
        self.__worksheet = self.__workbook.get_sheet_by_name(title)
        exec_count = node.get_execution_count()

        for current_fuse in fusions:
            for key, value in current_fuse.items():
                content = node.get_inner_content().replace("\l\t", "\n")

                if highlight_fuse and rich_text_is_available:

                    key = str(key) + "\n"
                    value = str(value) + "\n"
                    key_index = content.index(key)
                    value_index = content.index(value)

                    rich_string = CellRichText(
                        content[: key_index],
                        TextBlock(red_font, key),
                        content[key_index + len(key): value_index],
                        TextBlock(red_font, value),
                        content[value_index + len(value):]
                    )

                    content = rich_string

                input_out = f"{key} \n{value}"
                insn_profit = int(int(exec_count) / len(node))
                row = [func_name, content, input_out, int(exec_count), insn_profit]
                self.__worksheet.__rows.append(row)


    def dump(self, row_id: int) -> NoReturn:
        for worksheet in self.__workbook.get_sheet_names():
            self.__worksheet = self.__workbook.get_sheet_by_name(worksheet)
            self.__worksheet.__rows.sort(key=lambda k: k[row_id], reverse=True)

            for row in self.__worksheet.__rows:
                self.__worksheet.append(row)
                self.__worksheet.row_dimensions[self.__worksheet.max_row].height = 30
                self.__worksheet.row_dimensions[self.__worksheet.max_row].alignment = \
                    Alignment(vertical="top")
                self.__worksheet.append([" ", " ", " ", " "])

        self.__workbook.save(self.__xlsx_file)


    def append(self, graph: FlowGraph,
               func_name: str) -> NoReturn:

        for node in graph.nodes:
            if node.is_singleton:
                content = node.get_inner_content().replace("\l\t", "\n")
                bb_address_label = f"{node.get_address().strip(':')} " \
                                   f"({node.get_label().strip(':')})"
                exec_count = node.get_execution_count()
                row = [func_name, bb_address_label, content, exec_count]

                self.__worksheet.__rows.append(row)
